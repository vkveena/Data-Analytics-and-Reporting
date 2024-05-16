#DataCreationForGeneralInformation
import openpyxl
import random

c = "C:\\Users\\vardh\\OneDrive\\Desktop\\TableauProject\\InPatient_CMS\\Inpatient_Rehabilitation_Facility-General_Information_Mar2023.xlsx"
#c = "C:\\Users\\vardh\\OneDrive\Desktop\\TableauProject\\InPatient_CMS\\Sample.xlsx"

# Load the existing workbook
workbook = openpyxl.load_workbook(c)

# Select the worksheet where you want to add the new columns
worksheet = workbook["Inpatient_Rehabilitation_Facili"]
#worksheet = workbook["Sample"]

#Specify the range for the random values - Rating
lower_bound = 1
upper_bound = 10

#Specify the range for the random values - Revenue
lower_bound1 = 500000
upper_bound1 = 1100000

# Specify the range for the random values - Profit %
lower_bound2 = -20
upper_bound2 = 20

print(worksheet.max_row)
print(worksheet.max_column)

b = worksheet.max_column

# random_value1 = random.randint(lower_bound1, upper_bound1)
# random_value2 = random.randint(lower_bound2, upper_bound2)
# ProfitCal = (random_value2/100)*random_value1
# print (random_value1)
# print (random_value2)
# print (random_value2/100)
# print (ProfitCal)

# # Generate random values for each row within the range
for row in range(2, worksheet.max_row+1):
    random_value = random.randint(lower_bound, upper_bound)
    random_value1 = random.uniform(lower_bound1, upper_bound1)
    random_value2 = random.randint(lower_bound2,   upper_bound2)
    worksheet.cell(row=row, column= b+1).value = random_value
    worksheet.cell(row=row, column= b+2).value = round(random_value1)
    
    Ownership = worksheet.cell(row=row, column=b-1).value
    if  Ownership in ["For profit"] : worksheet.cell(row=row, column= b+3).value = random_value2
    else : worksheet.cell(row=row, column= b).value = "NA" 
        

# Add headers for the new columns
worksheet.cell(row=1, column=b+1).value = "Rating"
worksheet.cell(row=1, column=b+2).value = "Gross Annual Revenue"
worksheet.cell(row=1, column=b+3).value = "Profit Percentage"


# Save the changes to the workbook
workbook.save(c)
